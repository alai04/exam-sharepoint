import json
import io

from dotenv import dotenv_values
import yfinance as yf
from openpyxl import load_workbook
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File

config = dotenv_values()
ctx = ClientContext(config['SITE_URL'])
ctx.with_user_credentials(config['USERNAME'], config['PASSWORD'])
site = ctx.site.get().execute_query()
print(site.url)
list = ctx.web.lists.get_by_title('Tracked companies')

with open(config['TICKERS_FILE']) as tks_file:
    tks_cfg = json.load(tks_file)
print(tks_cfg)


def getTargetPrice(ticker):
    response = File.open_binary(ctx, config['SP_FILE_PREFIX']+ticker['model'])
    bytes_file_obj = io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0)
    wb = load_workbook(bytes_file_obj, read_only=True, data_only=True)
    ws = wb[ticker['sheet']]
    return float(ws[ticker['tpLocation']].value)


def getCurrentPriceAndName(ticker):
    try:
        tk = yf.Ticker(ticker)
        return tk.info['currentPrice'], tk.info['shortName']
    except:
        return None, None


def update1Ticker(tk):
    print(tk)
    if len(tk['name']) == 0:
        return

    filter_text = "Title eq '{0}'".format(tk['name'])
    payload = {
        'Title': tk['name'],
        'field_2': '{0:.2f}'.format(tk['cp']),
        'field_3': '{0:.2f}'.format(tk['tp']),
        'field_4': '{0:.2f}'.format(tk['tp']/tk['cp']-1.0),
    }
    items = list.items.filter(filter_text).get().execute_query()
    if len(items) == 0:
        list.add_item(payload).execute_query()
    else:
        items[0].validate_update_list_item(payload).execute_query()


def updateSPList(tks):
    items = list.items.get().execute_query()
    print('Before update, there are {0} items in list.'.format(len(items)))

    for tk in tks:
        update1Ticker(tk)

    items = list.items.get().execute_query()
    print('After update, there are {0} items in list.'.format(len(items)))
    return


if __name__ == '__main__':
    tks = []
    for tk in tks_cfg['tickers']:
        tp = getTargetPrice(tk)
        if not tp:
            continue

        cp, name = getCurrentPriceAndName(tk['ticker'])
        if not cp:
            continue

        tk['tp'] = tp
        tk['cp'] = cp
        tk['name'] = name
        print('Get price of {0}: ${1:.2f}, ${2:.2f}'.format(
            tk['ticker'], tp, cp))
        tks.append(tk)

    updateSPList(tks)
